import openpyxl

book = openpyxl.load_workbook("testdata.xlsx")
sheet = book.active

# Read
cell = sheet.cell(row=1, column=2)
print(cell.value)

# Write
sheet.cell(row=2, column=2).value = "Selenium"
print(sheet.cell(row=2, column=2).value)

# Max Rows
print(sheet.max_row)

# Max Column
print(sheet.max_column)

print(sheet['A5'].value)

dict = {}

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "testcase2":
        for j in range(2, sheet.max_column+1):
            print(f"Element at Row {i} Column {i} is {sheet.cell(row=i, column=j).value}")
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(dict)