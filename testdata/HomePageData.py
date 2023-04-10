import openpyxl


class HomePageData:

    test_homepage_tup = [("Selenium", "hello@selenium.com", "Sel123", "Male", "01012022"),
                         ("Automation", "hello@automation.com", "Aut123", "Female", "31012022")]
    test_homepage_dict = [{"Name": "Selenium", "E-Mail": "hello@selenium.com", "Password": "Sel123",
                           "Gender": "Male", "DOB": "01012022"},
                          {"Name": "Automation", "E-Mail": "hello@automation.com", "Password": "Aut123",
                           "Gender": "Female", "DOB": "01012022"}]

    @staticmethod
    def getTestData(*testcasenames):

        book = openpyxl.load_workbook("E:\\Skillsets\\Automation_Testing\\Projects\\Python_Selenium_Framework\\testdata\\testdata.xlsx")
        sheet = book.active

        dicti = {}
        listi = []
        for testcasename in testcasenames:
            for i in range(1, sheet.max_row + 1):
                if sheet.cell(row=i, column=1).value == testcasename:
                    for j in range(2, sheet.max_column + 1):
                        dicti[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            listi.append(dicti)

        return listi